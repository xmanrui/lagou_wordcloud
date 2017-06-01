# -*- coding-8 -*-
from openpyxl import load_workbook
from common import get_job_type_from_position_info_xlsx


def get_common(xlsx_file, row_name):
    job_type = get_job_type_from_position_info_xlsx(xlsx_file)
    wb = load_workbook(xlsx_file)
    ws = wb.get_sheet_by_name(job_type)
    rows = len(list(ws.rows))
    row_format = row_name + '%s'
    content = [ws[row_format % i].value for i in range(1, rows+1)]

    return content


def get_company_ids(xlsx_file):
    ids_list = get_common(xlsx_file, 'L')  # L列是companyID
    ids_list = filter(lambda x: x, ids_list)
    return list(ids_list)


def get_salary_list(xlsx_file):
    salary_list = get_common(xlsx_file, 'E')  # E列是salary
    salary_list = filter(lambda x: x, salary_list)  # 过滤掉None
    return list(salary_list)


def get_district(xlsx_file):
    district_list = get_common(xlsx_file, 'I')
    district_list = filter(lambda x: x, district_list)  # 过滤掉None
    return list(district_list)


def get_business_zones(xlsx_file):
    zones = get_common(xlsx_file, 'K')
    zones = filter(lambda x: x, zones)
    return list(zones)


def get_industry_field(xlsx_file):
    fields = get_common(xlsx_file, 'D')
    fields = filter(lambda x: x, fields)
    fields = list(fields)
    result = list()
    # 把 "移动互联网,生活服务" 拆成[移动互联网,生活服务]
    [result.extend(item.split(',')) for item in fields]

    return result


def test_get_company_id():
    test_file = './xlsx_file/python_position_info.xlsx'
    a = get_company_ids(test_file)
    print(a)


def test_get_district():
    test_file = './xlsx_file/python_position_info.xlsx'
    a = get_district(test_file)
    print(a)


def test_get_business_zones():
    test_file = './xlsx_file/python_position_info.xlsx'
    a = get_business_zones(test_file)
    print(a)


def test_get_industry_field():
    test_file = './xlsx_file/python_position_info.xlsx'
    a = get_industry_field(test_file)
    print(a)

if __name__ == '__main__':
    # test_get_company_id()
    # test_get_district()
    # test_get_business_zones()
    test_get_industry_field()
